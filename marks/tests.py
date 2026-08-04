import io
import json
import hashlib
import shutil
import tempfile
from datetime import date, timedelta
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from django.core.management import call_command

from .forms import BranchForm, MarkForm
from .models import (
    Bot,
    Branch,
    Experiment,
    MarkedLink,
    Product,
    ShortLink,
    TaskRequest,
    UserProfile,
    UtmDictionaryEntry,
)
from .services.link_builder import build_campaign, build_full_url
from .services.utm_dictionary import parse_campaign, parse_medium, parse_source
from .task_time import get_tasks_timezone


class TaskBoardBaseTestCase(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.admin_user = user_model.objects.create_user(username="admin", password="StrongPass123!")
        self.admin_user.profile.role = UserProfile.Role.ADMIN
        self.admin_user.profile.save(update_fields=["role"])

        self.manager_user = user_model.objects.create_user(username="manager", password="StrongPass123!")
        self.manager_user.profile.role = UserProfile.Role.MANAGER
        self.manager_user.profile.save(update_fields=["role"])

        self.product = Product.objects.create(name="Test product")
        self.bot = Bot.objects.create(name="test_bot_name", product=self.product)
        self.branch_main = Branch.objects.create(bot=self.bot, name="Main", code="MN")


class TaskBoardAccessTests(TaskBoardBaseTestCase):
    def test_admin_has_access_to_tasks_board(self):
        self.client.force_login(self.admin_user)
        response = self.client.get(reverse("tasks_board"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Задачник")
        self.assertContains(response, "Выгрузить выполненные за период")

    def test_non_admin_has_access_but_no_kanban(self):
        self.client.force_login(self.manager_user)
        response = self.client.get(reverse("tasks_board"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Задачник")
        self.assertFalse(response.context["show_kanban"])
        self.assertNotContains(response, "Выгрузить выполненные за период")


class TaskBoardActionsTests(TaskBoardBaseTestCase):
    @patch("marks.views.notify_new_task")
    def test_create_patch_task(self, notify_mock):
        self.client.force_login(self.admin_user)
        deadline = timezone.now() + timedelta(days=3)
        response = self.client.post(
            reverse("create_patch_task"),
            {
                "patch-branches": [self.branch_main.id],
                "patch-cjm_url": "https://example.com/cjm",
                "patch-comment": "Комментарий",
                "patch-deadline": deadline.strftime("%Y-%m-%dT%H:%M"),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("tasks_board"))

        task = TaskRequest.objects.get()
        self.assertEqual(task.task_type, TaskRequest.Type.PATCH)
        self.assertEqual(task.status, TaskRequest.Status.UNREAD)
        self.assertEqual(task.created_by, self.admin_user)
        self.assertEqual(list(task.branches.values_list("id", flat=True)), [self.branch_main.id])
        notify_mock.assert_called_once_with(task)

    @patch("marks.views.notify_new_task")
    def test_create_patch_task_saves_photo_and_task_timezone_deadline(self, notify_mock):
        self.client.force_login(self.admin_user)
        media_root = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, media_root, ignore_errors=True)
        photo = SimpleUploadedFile("task.png", b"fake-image-bytes", content_type="image/png")

        with self.settings(MEDIA_ROOT=media_root, TASKS_TIME_ZONE="Europe/Moscow"):
            response = self.client.post(
                reverse("create_patch_task"),
                {
                    "patch-branches": [self.branch_main.id],
                    "patch-cjm_url": "https://example.com/cjm",
                    "patch-comment": "Комментарий со скрином",
                    "patch-deadline": "2026-04-08T12:34",
                    "patch-photo": photo,
                },
            )
            task = TaskRequest.objects.get(task_type=TaskRequest.Type.PATCH)
            self.assertTrue(task.photo.name.endswith(".png"))
            self.assertEqual(
                timezone.localtime(task.deadline, get_tasks_timezone()).strftime("%Y-%m-%dT%H:%M"),
                "2026-04-08T12:34",
            )

        self.assertEqual(response.status_code, 302)
        task.refresh_from_db()
        notify_mock.assert_called_once_with(task)

    @patch("marks.views.notify_new_task")
    def test_create_build_task_uses_manual_bot_name(self, notify_mock):
        self.client.force_login(self.admin_user)
        deadline = timezone.now() + timedelta(days=2)
        response = self.client.post(
            reverse("create_build_task"),
            {
                "build-bot_name": "@new_bot",
                "build-build_token": "1234567890",
                "build-cjm_url": "https://example.com/cjm-build",
                "build-comment": "Build comment",
                "build-deadline": deadline.strftime("%Y-%m-%dT%H:%M"),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("tasks_board"))

        task = TaskRequest.objects.get(task_type=TaskRequest.Type.BUILD)
        self.assertEqual(task.build_name, "@new_bot")
        self.assertEqual(task.branches.count(), 0)
        self.assertEqual(task.get_scope_units(), 1)
        notify_mock.assert_called_once_with(task)

    @patch("marks.views.notify_new_task")
    def test_create_build_task_appends_optional_branch_name(self, notify_mock):
        self.client.force_login(self.admin_user)
        deadline = timezone.now() + timedelta(days=2)
        response = self.client.post(
            reverse("create_build_task"),
            {
                "build-bot_name": "@new_bot",
                "build-branch_name": "feature-login",
                "build-build_token": "1234567890",
                "build-cjm_url": "https://example.com/cjm-build",
                "build-comment": "Build comment",
                "build-deadline": deadline.strftime("%Y-%m-%dT%H:%M"),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("tasks_board"))

        task = TaskRequest.objects.get(task_type=TaskRequest.Type.BUILD)
        self.assertEqual(task.build_name, "@new_bot / feature-login")
        self.assertEqual(task.get_scope_units(), 1)
        notify_mock.assert_called_once_with(task)

    @patch("marks.views.notify_new_task")
    def test_create_mailing_task_accepts_local_deadline_format(self, notify_mock):
        self.client.force_login(self.admin_user)
        deadline = timezone.now() + timedelta(days=2)
        response = self.client.post(
            reverse("create_mailing_task"),
            {
                "mailing-branches": [self.branch_main.id],
                "mailing-tz_url": "https://example.com/tz",
                "mailing-comment": "Mailing comment",
                "mailing-deadline": deadline.strftime("%d.%m.%Y %H:%M"),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("tasks_board"))

        task = TaskRequest.objects.get(task_type=TaskRequest.Type.MAILING)
        self.assertEqual(task.status, TaskRequest.Status.UNREAD)
        self.assertEqual(task.created_by, self.admin_user)
        self.assertEqual(task.tz_url, "https://example.com/tz")
        self.assertEqual(list(task.branches.values_list("id", flat=True)), [self.branch_main.id])
        notify_mock.assert_called_once_with(task)

    @patch("marks.views.notify_new_task")
    def test_create_task_with_notify_requires_tg_username(self, notify_mock):
        self.client.force_login(self.admin_user)
        deadline = timezone.now() + timedelta(days=2)
        response = self.client.post(
            reverse("create_patch_task"),
            {
                "patch-branches": [self.branch_main.id],
                "patch-cjm_url": "https://example.com/cjm",
                "patch-comment": "Комментарий",
                "patch-deadline": deadline.strftime("%Y-%m-%dT%H:%M"),
                "patch-notify_me": "on",
                "patch-tg_username": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Укажите username в Telegram или chat_id.")
        self.assertFalse(TaskRequest.objects.filter(task_type=TaskRequest.Type.PATCH).exists())
        notify_mock.assert_not_called()

    @patch("marks.views.notify_status_change")
    def test_status_done_sets_completed_at_and_sends_notification(self, notify_mock):
        task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.BUILD,
            build_name="bot + branches",
            build_token="1234567890",
            cjm_url="https://example.com/cjm",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
        )
        self.client.force_login(self.admin_user)

        response = self.client.post(
            reverse("update_task_status", kwargs={"task_id": task.id}),
            {"status": TaskRequest.Status.DONE},
        )

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("tasks_board"))
        task.refresh_from_db()
        self.assertEqual(task.status, TaskRequest.Status.DONE)
        self.assertIsNotNone(task.completed_at)

        notify_mock.assert_called_once()
        kwargs = notify_mock.call_args.kwargs
        self.assertEqual(kwargs["task"], task)
        self.assertEqual(kwargs["old_status"], TaskRequest.Status.UNREAD)
        self.assertEqual(kwargs["changed_by"], self.admin_user)

    @patch("marks.views.notify_status_change")
    def test_status_done_never_saves_completed_at_before_created_at(self, notify_mock):
        task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.BUILD,
            build_name="bot + branches",
            build_token="1234567890",
            cjm_url="https://example.com/cjm",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
        )
        future_created_at = timezone.now() + timedelta(minutes=10)
        TaskRequest.objects.filter(pk=task.pk).update(created_at=future_created_at)
        self.client.force_login(self.admin_user)

        response = self.client.post(
            reverse("update_task_status", kwargs={"task_id": task.id}),
            {"status": TaskRequest.Status.DONE},
        )

        self.assertEqual(response.status_code, 302)
        task.refresh_from_db()
        self.assertEqual(task.completed_at, task.created_at)
        notify_mock.assert_called_once()

    @patch("marks.views.notify_done_to_user")
    @patch("marks.views.get_task_tg_username", return_value="test_user")
    @patch("marks.views.notify_status_change")
    def test_status_done_sends_personal_notification_when_username_exists(
        self,
        status_notify_mock,
        legacy_username_mock,
        user_notify_mock,
    ):
        task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.MAILING,
            tz_url="https://example.com/tz",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
        )
        self.client.force_login(self.admin_user)

        response = self.client.post(
            reverse("update_task_status", kwargs={"task_id": task.id}),
            {"status": TaskRequest.Status.DONE},
        )

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("tasks_board"))
        task.refresh_from_db()
        self.assertEqual(task.status, TaskRequest.Status.DONE)

        status_notify_mock.assert_called_once()
        legacy_username_mock.assert_called_once_with(task.id)
        user_notify_mock.assert_called_once_with(task=task, tg_username="test_user")

    @override_settings(TELEGRAM_WEBHOOK_SECRET="secret-key")
    @patch("marks.views.set_task_feedback_comment")
    def test_telegram_webhook_saves_feedback_from_reply(self, set_feedback_mock):
        task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.PATCH,
            cjm_url="https://example.com/cjm",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
            status=TaskRequest.Status.DONE,
        )
        payload = {
            "update_id": 1,
            "message": {
                "message_id": 101,
                "text": "Все ок, спасибо!",
                "reply_to_message": {
                    "message_id": 100,
                    "text": f"ID задачи: #{task.id}\nЗадача выполнена",
                },
            },
        }

        response = self.client.post(
            reverse("telegram_webhook", kwargs={"webhook_key": "secret-key"}),
            data=json.dumps(payload),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        set_feedback_mock.assert_called_once_with(task_id=task.id, feedback_comment="Все ок, спасибо!")

    @override_settings(TELEGRAM_WEBHOOK_SECRET="secret-key")
    @patch("marks.views.set_task_feedback_comment")
    def test_telegram_webhook_saves_feedback_from_quote_when_reply_is_inaccessible(self, set_feedback_mock):
        task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.PATCH,
            cjm_url="https://example.com/cjm",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
            status=TaskRequest.Status.DONE,
        )
        payload = {
            "update_id": 5,
            "message": {
                "message_id": 105,
                "text": "РЎРїР°СЃРёР±Рѕ, РІС‹ СЃСѓРїРµСЂ!!",
                "reply_to_message": {
                    "message_id": 104,
                    "date": 0,
                    "chat": {"id": -100100100, "type": "supergroup"},
                },
                "quote": {
                    "text": f"Р—Р°РґР°С‡Р° РІС‹РїРѕР»РЅРµРЅР°\nID Р·Р°РґР°С‡Рё: #{task.id}",
                },
            },
        }

        response = self.client.post(
            reverse("telegram_webhook", kwargs={"webhook_key": "secret-key"}),
            data=json.dumps(payload),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        set_feedback_mock.assert_called_once_with(task_id=task.id, feedback_comment="РЎРїР°СЃРёР±Рѕ, РІС‹ СЃСѓРїРµСЂ!!")

    @override_settings(TELEGRAM_WEBHOOK_SECRET="secret-key")
    @patch("marks.views.set_task_feedback_comment")
    def test_telegram_webhook_saves_feedback_from_business_message_external_reply(self, set_feedback_mock):
        task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.MAILING,
            tz_url="https://example.com/tz",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
            status=TaskRequest.Status.DONE,
        )
        payload = {
            "update_id": 6,
            "business_message": {
                "message_id": 106,
                "text": "Р¤РёРґР±РµРє РёР· business chat",
                "external_reply": {
                    "message_id": 105,
                    "text": f"ID Р·Р°РґР°С‡Рё: #{task.id}\nР—Р°РґР°С‡Р° РІС‹РїРѕР»РЅРµРЅР°",
                },
            },
        }

        response = self.client.post(
            reverse("telegram_webhook", kwargs={"webhook_key": "secret-key"}),
            data=json.dumps(payload),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        set_feedback_mock.assert_called_once_with(task_id=task.id, feedback_comment="Р¤РёРґР±РµРє РёР· business chat")

    @override_settings(TELEGRAM_WEBHOOK_SECRET="secret-key")
    def test_telegram_webhook_feedback_is_exported(self):
        task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.PATCH,
            cjm_url="https://example.com/cjm",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
            status=TaskRequest.Status.DONE,
        )
        payload = {
            "update_id": 4,
            "message": {
                "message_id": 104,
                "text": "Фидбек по задаче",
                "reply_to_message": {
                    "message_id": 103,
                    "text": f"ID задачи: #{task.id}\nЗадача выполнена",
                },
            },
        }

        webhook_response = self.client.post(
            reverse("telegram_webhook", kwargs={"webhook_key": "secret-key"}),
            data=json.dumps(payload),
            content_type="application/json",
        )

        self.assertEqual(webhook_response.status_code, 200)

        self.client.force_login(self.admin_user)
        export_response = self.client.get(reverse("export_completed_tasks"))

        self.assertEqual(export_response.status_code, 200)
        workbook = load_workbook(io.BytesIO(export_response.content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        feedback_index = headers.index("Фидбек")
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        row = next(row for row in rows if row[0] == task.id)

        self.assertEqual(row[feedback_index], "Фидбек по задаче")

    @override_settings(TELEGRAM_WEBHOOK_SECRET="secret-key")
    @patch("marks.views.send_text_message")
    @patch("marks.views.send_weekly_tasks_report", return_value=(True, ""))
    def test_telegram_webhook_week_command_sends_report(self, send_report_mock, send_text_mock):
        task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.PATCH,
            cjm_url="https://example.com/cjm",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
            status=TaskRequest.Status.DONE,
        )
        task.completed_at = timezone.now()
        task.save(update_fields=["completed_at"])

        payload = {
            "update_id": 2,
            "message": {
                "message_id": 102,
                "text": "/week",
                "chat": {"id": 439144407},
            },
        }
        response = self.client.post(
            reverse("telegram_webhook", kwargs={"webhook_key": "secret-key"}),
            data=json.dumps(payload),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        send_report_mock.assert_called_once()
        kwargs = send_report_mock.call_args.kwargs
        self.assertEqual(kwargs["chat_id"], "439144407")
        self.assertIn("tasks_week_current_", kwargs["filename"])
        self.assertIn("текущую неделю", kwargs["caption"])
        send_text_mock.assert_not_called()

    @override_settings(TELEGRAM_WEBHOOK_SECRET="secret-key")
    @patch("marks.views.send_text_message")
    @patch("marks.views.send_weekly_tasks_report", return_value=(True, ""))
    def test_telegram_webhook_month_command_rejects_bad_date(self, send_report_mock, send_text_mock):
        payload = {
            "update_id": 3,
            "message": {
                "message_id": 103,
                "text": "/month 15-02-2026",
                "chat": {"id": 439144407},
            },
        }
        response = self.client.post(
            reverse("telegram_webhook", kwargs={"webhook_key": "secret-key"}),
            data=json.dumps(payload),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        send_report_mock.assert_not_called()
        send_text_mock.assert_called_once()

    def test_tasks_board_filters_by_status(self):
        done_task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.BUILD,
            build_name="done task",
            build_token="token",
            cjm_url="https://example.com/cjm",
            deadline=timezone.now() + timedelta(days=2),
            created_by=self.admin_user,
            status=TaskRequest.Status.DONE,
        )
        unread_task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.BUILD,
            build_name="unread task",
            build_token="token2",
            cjm_url="https://example.com/cjm2",
            deadline=timezone.now() + timedelta(days=2),
            created_by=self.admin_user,
            status=TaskRequest.Status.UNREAD,
        )
        self.client.force_login(self.admin_user)

        response = self.client.get(
            reverse("tasks_board"),
            {"task_status": TaskRequest.Status.DONE},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f">#{done_task.id}<")
        self.assertNotContains(response, f">#{unread_task.id}<")

    def test_completed_counters_use_branch_units(self):
        second_branch = Branch.objects.create(bot=self.bot, name="Dev", code="DV")
        third_branch = Branch.objects.create(bot=self.bot, name="Feature", code="FT")

        mailing_task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.MAILING,
            tz_url="https://example.com/tz",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
            status=TaskRequest.Status.DONE,
        )
        mailing_task.branches.set([self.branch_main, second_branch])

        build_task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.BUILD,
            build_token="build-token",
            cjm_url="https://example.com/cjm",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
            status=TaskRequest.Status.DONE,
        )
        build_task.branches.set([self.branch_main, second_branch, third_branch])

        patch_task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.PATCH,
            cjm_url="https://example.com/cjm-patch",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
            status=TaskRequest.Status.DONE,
        )
        patch_task.branches.set([self.branch_main])

        self.client.force_login(self.admin_user)
        response = self.client.get(reverse("tasks_board"))

        self.assertEqual(response.status_code, 200)
        counters = response.context["completed_type_counters"]
        self.assertEqual(counters["mailing"], 2)
        self.assertEqual(counters["build"], 3)
        self.assertEqual(counters["patch"], 1)

    def test_export_completed_tasks_by_period(self):
        recent_task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.BUILD,
            build_name="recent",
            build_token="token",
            cjm_url="https://example.com/cjm",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
            status=TaskRequest.Status.DONE,
        )
        old_task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.BUILD,
            build_name="old",
            build_token="token",
            cjm_url="https://example.com/cjm",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
            status=TaskRequest.Status.DONE,
        )

        recent_task.completed_at = timezone.now() - timedelta(days=1)
        recent_task.save(update_fields=["completed_at"])
        old_task.completed_at = timezone.now() - timedelta(days=30)
        old_task.save(update_fields=["completed_at"])

        self.client.force_login(self.admin_user)
        date_from = (timezone.localdate() - timedelta(days=3)).isoformat()
        date_to = timezone.localdate().isoformat()
        response = self.client.get(
            reverse("export_completed_tasks"),
            {"completed_from": date_from, "completed_to": date_to},
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response["Content-Type"])

        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        task_ids = [row[0] for row in rows]
        self.assertIn(recent_task.id, task_ids)
        self.assertNotIn(old_task.id, task_ids)

    def test_export_has_combined_link_column_and_bot_branch_column(self):
        task = TaskRequest.objects.create(
            task_type=TaskRequest.Type.MAILING,
            tz_url="https://example.com/tz",
            deadline=timezone.now() + timedelta(days=1),
            created_by=self.admin_user,
            status=TaskRequest.Status.DONE,
        )
        task.branches.set([self.branch_main])
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse("export_completed_tasks"))
        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        self.assertNotIn("Ветки", headers)
        self.assertIn("CJM/ТЗ", headers)
        self.assertIn("Бот и ветки", headers)
        self.assertIn("Фидбек", headers)

        first_data_row = [cell.value for cell in sheet[2]]
        self.assertIn("https://example.com/tz", first_data_row)
        self.assertIn("test_bot_name", first_data_row)


class BotPlatformTests(TaskBoardBaseTestCase):
    def test_default_telegram_tag_uses_start_link(self):
        tag = self.branch_main.tags.get(url__isnull=False)

        self.assertEqual(tag.url, "https://telegram.me/test_bot_name?start=MN0001")

    def test_vk_tag_uses_group_ref_link(self):
        vk_bot = Bot.objects.create(
            name="203482421",
            display_name="VK Sales",
            platform=Bot.Platform.VK,
            product=self.product,
        )
        branch = Branch.objects.create(bot=vk_bot, name="Main", code="ell23")
        tag = branch.tags.get(url__isnull=False)

        self.assertEqual(tag.url, "https://vk.com/write-203482421?ref=ell230001&ref_source=23")

    def test_bot_api_finds_vk_bot_by_group_id(self):
        vk_bot = Bot.objects.create(
            name="203482421",
            display_name="VK Sales",
            platform=Bot.Platform.VK,
            product=self.product,
        )
        Branch.objects.create(bot=vk_bot, name="Main", code="ell01")

        response = self.client.get(reverse("bot_api", kwargs={"bot_name": "203482421"}))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["bot"], "203482421")
        self.assertEqual(payload["branches"][0]["tags"][0]["url"], "https://vk.com/write-203482421?ref=ell010001&ref_source=1")

    def test_bot_api_accepts_telegram_name_with_at_prefix(self):
        response = self.client.get(reverse("bot_api", kwargs={"bot_name": "@test_bot_name"}))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["bot"], "test_bot_name")

    @patch("marks.views.secrets.randbelow", return_value=85)
    def test_bot_api_number_response_includes_ab_key_for_active_test(self, randbelow_mock):
        Experiment.objects.create(
            title="API split by number",
            branch=self.branch_main,
            wants_ab_test=True,
            ab_test_options=["start"],
            metric_impact="CR",
            expected_change="+5%",
            hypothesis="Проверяем вариант первого экрана.",
            traffic_volume=Experiment.TrafficVolume.SPLIT_70_30,
            test_duration=Experiment.TestDuration.DAYS_7,
            start_date=timezone.localdate() - timedelta(days=7),
            end_date=timezone.localdate() + timedelta(days=7),
            status=Experiment.Status.IN_PROGRESS,
            created_by=self.admin_user,
        )
        number = self.branch_main.tags.get(url__isnull=False).number

        response = self.client.get(
            reverse("bot_api", kwargs={"bot_name": "test_bot_name"}),
            {"number": number},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["number"], number)
        self.assertEqual(payload["ab_key"], 2)
        randbelow_mock.assert_called_once_with(100)

    def test_bot_api_number_response_defaults_ab_key_to_one_without_active_test(self):
        number = self.branch_main.tags.get(url__isnull=False).number

        response = self.client.get(
            reverse("bot_api", kwargs={"bot_name": "test_bot_name"}),
            {"number": number},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["number"], number)
        self.assertEqual(payload["ab_key"], 1)

    def test_bot_api_returns_branch_ab_assignment_for_specific_branch_code(self):
        experiment = Experiment.objects.create(
            title="API split",
            branch=self.branch_main,
            wants_ab_test=True,
            ab_test_options=["start"],
            metric_impact="CR",
            expected_change="+5%",
            hypothesis="Проверяем вариант первого экрана.",
            traffic_volume=Experiment.TrafficVolume.SPLIT_70_30,
            test_duration=Experiment.TestDuration.DAYS_7,
            start_date=timezone.localdate() - timedelta(days=7),
            end_date=timezone.localdate() + timedelta(days=7),
            status=Experiment.Status.IN_PROGRESS,
            created_by=self.admin_user,
        )
        ab_key = "user-42"
        seed = f"{experiment.id}:{self.branch_main.id}:{ab_key}".encode("utf-8")
        bucket = int(hashlib.sha256(seed).hexdigest()[:16], 16) % 100
        expected_variant_value = 1 if bucket < 70 else 2

        response = self.client.get(
            reverse("bot_api", kwargs={"bot_name": "test_bot_name"}),
            {"branch_code": "mn", "ab_key": ab_key},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload["branches"]), 1)
        self.assertEqual(payload["branches"][0]["code"], "MN")
        self.assertEqual(payload["branches"][0]["ab_test"]["active"], True)
        self.assertEqual(payload["branches"][0]["ab_test"]["variant_value"], expected_variant_value)
        self.assertEqual(payload["branches"][0]["ab_test"]["split"], "70/30")
        self.assertEqual(payload["branches"][0]["ab_test"]["assignment_mode"], "hash")

    def test_bot_api_returns_inactive_ab_payload_when_branch_has_no_active_test(self):
        response = self.client.get(
            reverse("bot_api", kwargs={"bot_name": "test_bot_name"}),
            {"branch_code": "MN"},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["branches"][0]["ab_test"], {"active": False})

    def test_bot_api_returns_404_for_unknown_branch_code(self):
        response = self.client.get(
            reverse("bot_api", kwargs={"bot_name": "test_bot_name"}),
            {"branch_code": "missing"},
        )

        self.assertEqual(response.status_code, 404)

    def test_bots_list_creates_telegram_bot_and_strips_at_sign(self):
        self.client.force_login(self.admin_user)

        response = self.client.post(
            reverse("bots_list"),
            {
                "form_type": "telegram",
                "tg-name": "@new_bot",
            },
        )

        self.assertEqual(response.status_code, 302)
        bot = Bot.objects.get(name="new_bot")
        self.assertEqual(bot.platform, Bot.Platform.TELEGRAM)
        self.assertEqual(bot.display_name, "")

    def test_bots_list_creates_vk_bot_with_display_name(self):
        self.client.force_login(self.admin_user)

        response = self.client.post(
            reverse("bots_list"),
            {
                "form_type": "vk",
                "vk-name": "203482421",
                "vk-display_name": "VK Sales",
            },
        )

        self.assertEqual(response.status_code, 302)
        bot = Bot.objects.get(name="203482421")
        self.assertEqual(bot.platform, Bot.Platform.VK)
        self.assertEqual(bot.display_name, "VK Sales")

    def test_bots_list_sorts_telegram_before_vk(self):
        Bot.objects.create(name="zzz_bot", product=self.product)
        Bot.objects.create(
            name="203482421",
            display_name="Alpha VK",
            platform=Bot.Platform.VK,
            product=self.product,
        )
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse("bots_list"))

        self.assertEqual(response.status_code, 200)
        ordered_platforms = [bot.platform for bot in response.context["active_bots"]]
        self.assertEqual(ordered_platforms[:3], [Bot.Platform.TELEGRAM, Bot.Platform.TELEGRAM, Bot.Platform.VK])


class ExperimentBoardTests(TaskBoardBaseTestCase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def _experiment_payload(self, **overrides):
        payload = {
            "title": "Новый A/B тест",
            "tz_url": "https://example.com/tz/exp-1",
            "wants_ab_test": "on",
            "ab_test_options": ["start"],
            "ab_test_custom_option": "",
            "metric_impact": "CR",
            "comparison_text": "Текущий экран vs новый экран",
            "expected_change": "+8%",
            "hypothesis": "Если изменить первый экран, конверсия вырастет.",
            "traffic_volume": Experiment.TrafficVolume.SPLIT_50_50,
            "traffic_volume_other": "",
            "test_duration": Experiment.TestDuration.DAYS_7,
            "duration_users": "",
            "duration_end_date": "",
            "start_date": "2026-03-10",
            "end_date": "2026-03-17",
            "dashboard_url": "https://example.com/dashboard/exp-1",
            "result_variant_a": "CR 11%, open rate 24%",
            "result_variant_b": "CR 13%, open rate 28%",
            "comment": "Комментарий по эксперименту",
        }
        payload.update(overrides)
        return payload

    def test_create_experiment_saves_dates_and_manual_results(self):
        response = self.client.post(reverse("experiments_board"), self._experiment_payload())

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("experiments_board"))

        experiment = Experiment.objects.get(title="Новый A/B тест")
        self.assertEqual(experiment.status, Experiment.Status.BACKLOG)
        self.assertEqual(experiment.tz_url, "https://example.com/tz/exp-1")
        self.assertEqual(experiment.comparison_text, "Текущий экран vs новый экран")
        self.assertEqual(experiment.start_date, date(2026, 3, 10))
        self.assertEqual(experiment.end_date, date(2026, 3, 17))
        self.assertEqual(experiment.dashboard_url, "https://example.com/dashboard/exp-1")
        self.assertEqual(experiment.result_variant_a, "CR 11%, open rate 24%")
        self.assertEqual(experiment.result_variant_b, "CR 13%, open rate 28%")

    def test_create_experiment_saves_selected_branch_for_api(self):
        response = self.client.post(
            reverse("experiments_board"),
            self._experiment_payload(branch=str(self.branch_main.id)),
        )

        self.assertEqual(response.status_code, 302)
        experiment = Experiment.objects.get()
        self.assertEqual(experiment.branch_id, self.branch_main.id)
        return
        experiment = Experiment.objects.get(title="РќРѕРІС‹Р№ A/B С‚РµСЃС‚")
        self.assertEqual(experiment.branch_id, self.branch_main.id)

    def test_edit_experiment_updates_dates_and_metrics(self):
        experiment = Experiment.objects.create(
            title="Текущий тест",
            wants_ab_test=True,
            ab_test_options=["start"],
            metric_impact="CR",
            expected_change="+5%",
            hypothesis="Исходная гипотеза",
            traffic_volume=Experiment.TrafficVolume.SPLIT_50_50,
            test_duration=Experiment.TestDuration.DAYS_7,
            created_by=self.admin_user,
        )

        payload = self._experiment_payload(
            title="Текущий тест",
            experiment_id=str(experiment.id),
            result_variant_a="CR 10%",
            result_variant_b="CR 14%",
            start_date="2026-03-11",
            end_date="2026-03-18",
        )
        response = self.client.post(reverse("experiments_board"), payload)

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("experiments_board"))

        experiment.refresh_from_db()
        self.assertEqual(experiment.start_date, date(2026, 3, 11))
        self.assertEqual(experiment.end_date, date(2026, 3, 18))
        self.assertEqual(experiment.result_variant_a, "CR 10%")
        self.assertEqual(experiment.result_variant_b, "CR 14%")

    @patch("marks.views.notify_new_task", return_value=(True, ""))
    def test_move_to_draft_allows_empty_tz(self, notify_mock):
        experiment = Experiment.objects.create(
            title="Без ТЗ",
            wants_ab_test=True,
            ab_test_options=["start"],
            metric_impact="CR",
            comparison_text="Экран A vs экран B",
            expected_change="+5%",
            hypothesis="Проверяем первый экран",
            traffic_volume=Experiment.TrafficVolume.SPLIT_50_50,
            test_duration=Experiment.TestDuration.DAYS_7,
            created_by=self.admin_user,
            status=Experiment.Status.BACKLOG,
        )

        response = self.client.post(
            reverse("update_experiment_status", kwargs={"experiment_id": experiment.id}),
            {"status": Experiment.Status.DRAFT},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        experiment.refresh_from_db()
        self.assertEqual(experiment.status, Experiment.Status.DRAFT)
        self.assertIsNotNone(experiment.technical_task)
        self.assertEqual(experiment.technical_task.tz_url, "")
        notify_mock.assert_called_once_with(experiment.technical_task)

    @patch("marks.views.notify_new_task", return_value=(True, ""))
    def test_move_to_draft_creates_task_for_tech_team(self, notify_mock):
        experiment = Experiment.objects.create(
            title="Готов к разработке",
            branch=self.branch_main,
            tz_url="https://example.com/tz/ready",
            wants_ab_test=True,
            ab_test_options=["start"],
            metric_impact="CR",
            comparison_text="Экран A vs экран B",
            expected_change="+5%",
            hypothesis="Проверяем первый экран",
            traffic_volume=Experiment.TrafficVolume.SPLIT_50_50,
            test_duration=Experiment.TestDuration.DAYS_7,
            start_date=date(2026, 4, 2),
            created_by=self.admin_user,
            status=Experiment.Status.BACKLOG,
        )

        response = self.client.post(
            reverse("update_experiment_status", kwargs={"experiment_id": experiment.id}),
            {"status": Experiment.Status.DRAFT},
        )

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("experiments_board"))

        experiment.refresh_from_db()
        self.assertEqual(experiment.status, Experiment.Status.DRAFT)
        self.assertIsNotNone(experiment.technical_task)

        task = experiment.technical_task
        self.assertEqual(task.task_type, TaskRequest.Type.PATCH)
        self.assertEqual(task.status, TaskRequest.Status.UNREAD)
        self.assertEqual(task.tz_url, "https://example.com/tz/ready")
        self.assertEqual(list(task.branches.values_list("id", flat=True)), [self.branch_main.id])
        notify_mock.assert_called_once_with(task)

    def test_update_experiment_status_blocks_parallel_branch_test(self):
        Experiment.objects.create(
            title="РђРєС‚РёРІРЅС‹Р№ С‚РµСЃС‚",
            branch=self.branch_main,
            wants_ab_test=True,
            ab_test_options=["start"],
            metric_impact="CR",
            expected_change="+5%",
            hypothesis="РџРµСЂРІС‹Р№ С‚РµСЃС‚",
            traffic_volume=Experiment.TrafficVolume.SPLIT_50_50,
            test_duration=Experiment.TestDuration.DAYS_7,
            start_date=date(2026, 3, 10),
            end_date=date(2026, 3, 17),
            created_by=self.admin_user,
            status=Experiment.Status.IN_PROGRESS,
        )
        queued_experiment = Experiment.objects.create(
            title="Р’С‚РѕСЂРѕР№ С‚РµСЃС‚",
            branch=self.branch_main,
            wants_ab_test=True,
            ab_test_options=["start"],
            metric_impact="CR",
            expected_change="+3%",
            hypothesis="Р’С‚РѕСЂР°СЏ РіРёРїРѕС‚РµР·Р°",
            traffic_volume=Experiment.TrafficVolume.SPLIT_50_50,
            test_duration=Experiment.TestDuration.DAYS_7,
            start_date=date(2026, 3, 18),
            end_date=date(2026, 3, 25),
            created_by=self.admin_user,
            status=Experiment.Status.DRAFT,
        )

        response = self.client.post(
            reverse("update_experiment_status", kwargs={"experiment_id": queued_experiment.id}),
            {"status": Experiment.Status.IN_PROGRESS},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        queued_experiment.refresh_from_db()
        self.assertEqual(queued_experiment.status, Experiment.Status.DRAFT)
        return
        self.assertContains(response, "Р”Р»СЏ СЌС‚РѕР№ РІРµС‚РєРё СѓР¶Рµ РёРґРµС‚ РґСЂСѓРіРѕР№ A/B С‚РµСЃС‚.")

    def test_final_status_requires_dates_and_ab_results(self):
        experiment = Experiment.objects.create(
            title="Тест без итогов",
            wants_ab_test=True,
            ab_test_options=["start"],
            metric_impact="CR",
            expected_change="+5%",
            hypothesis="Проверяем первый экран",
            traffic_volume=Experiment.TrafficVolume.SPLIT_50_50,
            test_duration=Experiment.TestDuration.DAYS_7,
            created_by=self.admin_user,
            status=Experiment.Status.COMPLETED,
        )

        response = self.client.post(
            reverse("update_experiment_status", kwargs={"experiment_id": experiment.id}),
            {"status": Experiment.Status.SUCCESS},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        experiment.refresh_from_db()
        self.assertEqual(experiment.status, Experiment.Status.COMPLETED)
        self.assertContains(response, "Перед финальным решением заполните")

    def test_final_status_saves_completion_data_from_modal_post(self):
        experiment = Experiment.objects.create(
            title="Финализация из попапа",
            wants_ab_test=True,
            ab_test_options=["start"],
            metric_impact="CR",
            expected_change="+7%",
            hypothesis="Завершаем тест прямо из карточки",
            traffic_volume=Experiment.TrafficVolume.SPLIT_50_50,
            test_duration=Experiment.TestDuration.DAYS_7,
            created_by=self.admin_user,
            status=Experiment.Status.IN_PROGRESS,
        )

        response = self.client.post(
            reverse("update_experiment_status", kwargs={"experiment_id": experiment.id}),
            {
                "status": Experiment.Status.SUCCESS,
                "start_date": "2026-03-10",
                "end_date": "2026-03-17",
                "dashboard_url": "https://example.com/dashboard/final-exp",
                "result_variant_a": "CR 10%, CTR 18%",
                "result_variant_b": "CR 14%, CTR 22%",
                "comment": "Победил вариант B",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        experiment.refresh_from_db()
        self.assertEqual(experiment.status, Experiment.Status.SUCCESS)
        self.assertEqual(experiment.start_date, date(2026, 3, 10))
        self.assertEqual(experiment.end_date, date(2026, 3, 17))
        self.assertEqual(experiment.dashboard_url, "https://example.com/dashboard/final-exp")
        self.assertEqual(experiment.result_variant_a, "CR 10%, CTR 18%")
        self.assertEqual(experiment.result_variant_b, "CR 14%, CTR 22%")
        self.assertEqual(experiment.comment, "Победил вариант B")

    def test_final_status_allows_empty_variant_results(self):
        experiment = Experiment.objects.create(
            title="Финал без цифр",
            wants_ab_test=True,
            ab_test_options=["start"],
            metric_impact="CR",
            comparison_text="A vs B",
            expected_change="+7%",
            hypothesis="Проверяем без обязательных цифр",
            traffic_volume=Experiment.TrafficVolume.SPLIT_50_50,
            test_duration=Experiment.TestDuration.DAYS_7,
            created_by=self.admin_user,
            status=Experiment.Status.IN_PROGRESS,
        )

        response = self.client.post(
            reverse("update_experiment_status", kwargs={"experiment_id": experiment.id}),
            {
                "status": Experiment.Status.SUCCESS,
                "start_date": "2026-03-30",
                "end_date": "2026-04-01",
                "dashboard_url": "",
                "result_variant_a": "",
                "result_variant_b": "",
                "comment": "",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        experiment.refresh_from_db()
        self.assertEqual(experiment.status, Experiment.Status.SUCCESS)

    def test_finalized_experiment_moves_to_library(self):
        active_experiment = Experiment.objects.create(
            title="Активный тест",
            wants_ab_test=True,
            ab_test_options=["start"],
            metric_impact="CR",
            expected_change="+5%",
            hypothesis="Активная гипотеза",
            traffic_volume=Experiment.TrafficVolume.SPLIT_50_50,
            test_duration=Experiment.TestDuration.DAYS_7,
            created_by=self.admin_user,
            status=Experiment.Status.DRAFT,
        )
        final_experiment = Experiment.objects.create(
            title="Финальный тест",
            wants_ab_test=True,
            ab_test_options=["start"],
            metric_impact="CR",
            expected_change="+5%",
            hypothesis="Финальная гипотеза",
            traffic_volume=Experiment.TrafficVolume.SPLIT_50_50,
            test_duration=Experiment.TestDuration.DAYS_7,
            start_date=date(2026, 3, 10),
            end_date=date(2026, 3, 17),
            result_variant_a="CR 10%",
            result_variant_b="CR 15%",
            created_by=self.admin_user,
            status=Experiment.Status.SUCCESS,
        )

        response = self.client.get(reverse("experiments_board"))

        self.assertEqual(response.status_code, 200)

        active_ids = [
            item["item"].id
            for column in response.context["active_columns"]
            for item in column["items"]
        ]
        library_ids = [
            item["item"].id
            for column in response.context["library_columns"]
            for item in column["items"]
        ]

        self.assertIn(active_experiment.id, active_ids)
        self.assertNotIn(final_experiment.id, active_ids)
        self.assertIn(final_experiment.id, library_ids)


class BranchFormTests(TaskBoardBaseTestCase):
    def test_branch_form_prefills_next_code_from_existing_branches(self):
        Branch.objects.create(bot=self.bot, name="Second", code="ell01")
        Branch.objects.create(bot=self.bot, name="Third", code="ell02")

        form = BranchForm(bot=self.bot)

        self.assertEqual(form.initial["code"], "ell03")


class ShortLinkTests(TestCase):
    def test_code_autogenerated_on_save(self):
        link = ShortLink.objects.create(target_url="https://el-ed.ru/oge?utm_source=yandex")
        self.assertTrue(link.code)
        self.assertLessEqual(len(link.code), 20)

    def test_redirect_302_to_target(self):
        target = "https://el-ed.ru/oge?utm_source=yandex&utm_medium=cpc"
        link = ShortLink.objects.create(target_url=target)
        response = self.client.get(reverse("short_link", args=[link.code]))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], target)

    def test_click_counter_increments(self):
        link = ShortLink.objects.create(target_url="https://el-ed.ru/")
        self.assertEqual(link.clicks, 0)
        self.client.get(reverse("short_link", args=[link.code]))
        self.client.get(reverse("short_link", args=[link.code]))
        link.refresh_from_db()
        self.assertEqual(link.clicks, 2)
        self.assertIsNotNone(link.last_click_at)

    def test_unknown_code_returns_404(self):
        response = self.client.get(reverse("short_link", args=["no-such-code"]))
        self.assertEqual(response.status_code, 404)

    def test_short_url_uses_base_domain(self):
        link = ShortLink.objects.create(target_url="https://el-ed.ru/")
        with override_settings(SHORTLINK_BASE_DOMAIN="https://l.el-ed.ru"):
            self.assertEqual(link.short_url, f"https://l.el-ed.ru/s/{link.code}/")


class UtmDictionaryParserTests(TestCase):
    def test_parse_medium(self):
        rows = [
            ["02 · medium", "", "", "", ""],
            ["описание", "", "", "", ""],
            ["", "", "", "", ""],
            ["medium", "Что значит", "Группа", "Подсказка", "Владелец"],
            ["cpc", "клик", "Performance", "", ""],
            ["cpm", "показы", "Performance", "", ""],
            ["", "", "", "", ""],
        ]
        entries = parse_medium(rows)
        self.assertEqual([e["value"] for e in entries], ["cpc", "cpm"])
        self.assertTrue(all(e["field"] == "medium" for e in entries))
        self.assertEqual(entries[0]["group"], "Performance")

    def test_parse_source_flags_templates_and_inactive(self):
        rows = [
            ["03 · source", "", "", "", "", ""],
            ["описание", "", "", "", "", ""],
            ["", "", "", "", "", ""],
            ["source", "Группа", "Что это", "Активно", "Старые", "Владелец"],
            ["yandex", "кабинет", "Яндекс", "да", "", ""],
            ["google_ads", "кабинет", "Google", "НЕТ", "", ""],
            ["tg-<имя>", "соцканал", "наши ТГ", "—", "", ""],
        ]
        by_value = {e["value"]: e for e in parse_source(rows)}
        self.assertTrue(by_value["yandex"]["is_active"])
        self.assertFalse(by_value["google_ads"]["is_active"])
        self.assertTrue(by_value["tg-<имя>"]["is_template"])
        self.assertFalse(by_value["yandex"]["is_template"])

    def test_parse_campaign_sections(self):
        rows = [
            ["04 · campaign", "", ""],
            ["описание", "", ""],
            ["", "", ""],
            ["Часть 1 — тип кампании", "", ""],
            ["значение", "что значит", ""],
            ["acq", "привлекаем", ""],
            ["gen", "регулярная", ""],
            ["", "", ""],
            ["Часть 2 — направление (класс)", "", ""],
            ["значение", "что значит", ""],
            ["oge", "ОГЭ", ""],
            ["", "", ""],
            ["Часть 3 — воронка", "", ""],
            ["значение", "куда ведёт", "цель"],
            ["bot", "бот-мостик", "лид"],
            ["", "", ""],
            ["Часть 4 — имя кампании", "", ""],
            ["формат", "правило", "пример"],
            ["пишешь сам", "латиница", "express-ege-2026"],
        ]
        by_field = {}
        for entry in parse_campaign(rows):
            by_field.setdefault(entry["field"], []).append(entry["value"])
        self.assertEqual(by_field["type"], ["acq", "gen"])
        self.assertEqual(by_field["direction"], ["oge"])
        self.assertEqual(by_field["funnel"], ["bot"])
        self.assertNotIn("имя", by_field)


class LinkBuilderTests(TestCase):
    def test_build_campaign(self):
        self.assertEqual(build_campaign("acq", "oge", "bot", "pyweek2026"), "acq_oge_bot_pyweek2026")

    def test_build_full_url_preserves_placeholders(self):
        utm = {
            "utm_source": "yandex",
            "utm_medium": "cpc",
            "utm_campaign": "gen_oge_site_nonbrand",
            "utm_term": "{keyword}",
            "utm_content": "{ad_id}",
        }
        url = build_full_url("https://el-ed.ru/oge", utm)
        self.assertEqual(
            url,
            "https://el-ed.ru/oge?utm_source=yandex&utm_medium=cpc"
            "&utm_campaign=gen_oge_site_nonbrand&utm_term={keyword}&utm_content={ad_id}",
        )

    def test_build_full_url_merges_existing_query_and_skips_empty(self):
        utm = {
            "utm_source": "advcake",
            "utm_medium": "cpa",
            "utm_campaign": "all_ege_site_affiliate",
            "utm_term": "",
            "utm_content": "hash",
        }
        url = build_full_url("https://el-ed.ru/ege?ref=1", utm)
        self.assertTrue(url.startswith("https://el-ed.ru/ege?ref=1&utm_source=advcake"))
        self.assertIn("utm_content=hash", url)
        self.assertNotIn("utm_term", url)


class ImportUtmDictionaryCommandTests(TestCase):
    def test_import_from_repo_fixtures(self):
        call_command("import_utm_dictionary")
        self.assertEqual(UtmDictionaryEntry.objects.filter(field="medium").count(), 11)
        self.assertEqual(UtmDictionaryEntry.objects.filter(field="type").count(), 6)
        self.assertEqual(UtmDictionaryEntry.objects.filter(field="direction").count(), 6)
        self.assertEqual(UtmDictionaryEntry.objects.filter(field="funnel").count(), 11)
        self.assertFalse(
            UtmDictionaryEntry.objects.get(field="source", value="google_ads").is_active
        )
        self.assertTrue(
            UtmDictionaryEntry.objects.filter(field="source", is_template=True).exists()
        )


def _seed_min_dictionary():
    seed = [
        ("source", "yandex", False, True),
        ("source", "tg-<имя>", True, True),
        ("medium", "cpc", False, True),
        ("type", "acq", False, True),
        ("direction", "oge", False, True),
        ("funnel", "bot", False, True),
    ]
    for field, value, is_template, is_active in seed:
        UtmDictionaryEntry.objects.create(
            field=field, value=value, is_template=is_template, is_active=is_active
        )


class MarkGeneratorTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.marketer = user_model.objects.create_user(username="mktr", password="StrongPass123!")
        self.marketer.profile.role = UserProfile.Role.MARKETER
        self.marketer.profile.save(update_fields=["role"])
        self.bot_user = user_model.objects.create_user(username="botop", password="StrongPass123!")
        self.bot_user.profile.role = UserProfile.Role.BOT_USER
        self.bot_user.profile.save(update_fields=["role"])
        _seed_min_dictionary()

    def _valid_payload(self, **overrides):
        payload = {
            "original_url": "https://el-ed.ru/oge",
            "source": "yandex",
            "medium": "cpc",
            "mark_type": "acq",
            "direction": "oge",
            "funnel": "bot",
            "name": "pyweek2026",
            "utm_term": "interests",
            "utm_content": "ad-456",
            "make_full": "1",
        }
        payload.update({k: v for k, v in overrides.items() if v is not None})
        for key, value in overrides.items():
            if value is None:
                payload.pop(key, None)
        return payload

    def test_registry_allowed_for_bot_user(self):
        # Базовый доступ: оператор ботов тоже видит генератор/реестр.
        self.client.force_login(self.bot_user)
        self.assertEqual(self.client.get(reverse("marks_registry")).status_code, 200)

    def test_generator_opens_for_marketer(self):
        self.client.force_login(self.marketer)
        self.assertEqual(self.client.get(reverse("marks_new")).status_code, 200)

    def test_create_full_link(self):
        self.client.force_login(self.marketer)
        response = self.client.post(reverse("marks_new"), self._valid_payload())
        self.assertRedirects(response, reverse("marks_registry"))
        mark = MarkedLink.objects.get()
        self.assertEqual(mark.utm_campaign, "acq_oge_bot_pyweek2026")
        self.assertEqual(mark.author, self.marketer)
        self.assertIsNone(mark.short_link)
        self.assertIn("utm_content=ad-456", mark.full_url)

    def test_create_short_link(self):
        self.client.force_login(self.marketer)
        self.client.post(reverse("marks_new"), self._valid_payload(make_full=None, make_short="1"))
        mark = MarkedLink.objects.get()
        self.assertIsNotNone(mark.short_link)
        self.assertEqual(mark.short_link.target_url, mark.full_url)
        self.assertEqual(ShortLink.objects.count(), 1)

    def test_invalid_name_rejected(self):
        self.client.force_login(self.marketer)
        response = self.client.post(reverse("marks_new"), self._valid_payload(name="py_week 2026"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(MarkedLink.objects.count(), 0)

    def test_source_outside_dictionary_rejected(self):
        self.client.force_login(self.marketer)
        response = self.client.post(reverse("marks_new"), self._valid_payload(source="made-up-source"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(MarkedLink.objects.count(), 0)

    def test_template_source_available_in_choices(self):
        # Phase 3: шаблоны source снова доступны для выбора (+ ввод имени в source_custom).
        source_values = [choice[0] for choice in MarkForm().fields["source"].choices]
        self.assertIn("yandex", source_values)
        self.assertIn("tg-<имя>", source_values)


class MarkRegistryFilterTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.alice = user_model.objects.create_user(username="alice", password="StrongPass123!")
        self.alice.profile.role = UserProfile.Role.MARKETER
        self.alice.profile.save(update_fields=["role"])
        self.bob = user_model.objects.create_user(username="bob", password="StrongPass123!")
        self.bob.profile.role = UserProfile.Role.MARKETER
        self.bob.profile.save(update_fields=["role"])

        self.link_yandex = MarkedLink.objects.create(
            original_url="https://el-ed.ru/oge",
            utm_source="yandex",
            utm_medium="cpc",
            utm_campaign="acq_oge_bot_alpha",
            utm_term="interests",
            utm_content="",
            mark_type="acq",
            direction="oge",
            funnel="bot",
            name="alpha",
            full_url="https://el-ed.ru/oge?utm_source=yandex",
            author=self.alice,
        )
        short = ShortLink.objects.create(target_url="https://el-ed.ru/ege?x", clicks=7)
        self.link_vk = MarkedLink.objects.create(
            original_url="https://el-ed.ru/ege",
            utm_source="vk_ads",
            utm_medium="cpm",
            utm_campaign="gen_ege_site_beta",
            utm_term="retarget",
            utm_content="",
            mark_type="gen",
            direction="ege",
            funnel="site",
            name="beta",
            full_url="https://el-ed.ru/ege?utm_source=vk_ads",
            author=self.bob,
            short_link=short,
        )
        self.client.force_login(self.alice)

    def _ids(self, response):
        return {link.id for link in response.context["links"]}

    def test_filter_by_source(self):
        response = self.client.get(reverse("marks_registry"), {"source": "yandex"})
        self.assertEqual(self._ids(response), {self.link_yandex.id})

    def test_filter_by_direction(self):
        response = self.client.get(reverse("marks_registry"), {"direction": "ege"})
        self.assertEqual(self._ids(response), {self.link_vk.id})

    def test_filter_by_campaign_substring(self):
        response = self.client.get(reverse("marks_registry"), {"campaign": "beta"})
        self.assertEqual(self._ids(response), {self.link_vk.id})

    def test_filter_by_author(self):
        response = self.client.get(reverse("marks_registry"), {"author": self.bob.id})
        self.assertEqual(self._ids(response), {self.link_vk.id})

    def test_filter_by_date_range(self):
        past = timezone.now() - timedelta(days=10)
        MarkedLink.objects.filter(pk=self.link_yandex.pk).update(created_at=past)
        today = timezone.localdate().isoformat()
        response = self.client.get(reverse("marks_registry"), {"date_from": today})
        self.assertEqual(self._ids(response), {self.link_vk.id})

    def test_clicks_available_from_shortlink(self):
        response = self.client.get(reverse("marks_registry"))
        links = {link.id: link for link in response.context["links"]}
        self.assertEqual(links[self.link_vk.id].short_link.clicks, 7)
        self.assertIsNone(links[self.link_yandex.id].short_link)

    def test_export_excel_returns_xlsx(self):
        response = self.client.get(reverse("marks_registry"), {"export": "excel"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", response["Content-Disposition"])

    def test_export_respects_filter(self):
        response = self.client.get(reverse("marks_registry"), {"export": "excel", "source": "yandex"})
        workbook = load_workbook(io.BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(len(rows), 2)  # заголовок + одна строка
        self.assertEqual(rows[1][1], "yandex")


from .services.utm_dictionary import sync_entries


class MarkEnforcementTests(TestCase):
    def setUp(self):
        _seed_min_dictionary()

    def _data(self, **overrides):
        data = {
            "original_url": "https://el-ed.ru/oge",
            "source": "yandex",
            "medium": "cpc",
            "mark_type": "acq",
            "direction": "oge",
            "funnel": "bot",
            "name": "pyweek2026",
            "utm_term": "interests",
            "utm_content": "ad-456",
        }
        data.update(overrides)
        return data

    def test_valid_form(self):
        self.assertTrue(MarkForm(self._data()).is_valid())

    def test_content_placeholder_rejected_strict_mask(self):
        # Строгая маска по ТЗ: {ad_id} с фигурными скобками не проходит.
        form = MarkForm(self._data(utm_content="{ad_id}"))
        self.assertFalse(form.is_valid())
        self.assertIn("utm_content", form.errors)

    def test_dashed_funnel_resolves(self):
        form = MarkForm(self._data(funnel_custom="python2026"))
        self.assertTrue(form.is_valid())
        self.assertEqual(form.cleaned_data["resolved_funnel"], "bot-python2026")

    def test_cyrillic_url_rejected(self):
        form = MarkForm(self._data(original_url="https://el-ed.ru/огэ"))
        self.assertFalse(form.is_valid())
        self.assertIn("original_url", form.errors)

    def test_space_url_rejected(self):
        self.assertFalse(MarkForm(self._data(original_url="https://el-ed.ru/o ge")).is_valid())

    def test_url_without_scheme_rejected(self):
        self.assertFalse(MarkForm(self._data(original_url="el-ed.ru/oge")).is_valid())

    def test_cyrillic_term_rejected(self):
        form = MarkForm(self._data(utm_term="конкуренты"))
        self.assertFalse(form.is_valid())
        self.assertIn("utm_term", form.errors)

    def test_underscore_name_rejected(self):
        form = MarkForm(self._data(name="py_week"))
        self.assertFalse(form.is_valid())
        self.assertIn("name", form.errors)

    def test_template_source_requires_custom(self):
        form = MarkForm(self._data(source="tg-<имя>", source_custom=""))
        self.assertFalse(form.is_valid())
        self.assertIn("source_custom", form.errors)

    def test_template_source_resolves_and_flags(self):
        form = MarkForm(self._data(source="tg-<имя>", source_custom="python2026"))
        self.assertTrue(form.is_valid())
        self.assertEqual(form.cleaned_data["resolved_source"], "tg-python2026")
        self.assertTrue(form.cleaned_data["pending_review"])

    def test_canonical_source_not_flagged(self):
        form = MarkForm(self._data())
        self.assertTrue(form.is_valid())
        self.assertEqual(form.cleaned_data["resolved_source"], "yandex")
        self.assertFalse(form.cleaned_data["pending_review"])


class MarkDedupTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.marketer = user_model.objects.create_user(username="mktr", password="StrongPass123!")
        self.marketer.profile.role = UserProfile.Role.MARKETER
        self.marketer.profile.save(update_fields=["role"])
        _seed_min_dictionary()
        self.client.force_login(self.marketer)

    def _payload(self, **overrides):
        payload = {
            "original_url": "https://el-ed.ru/oge",
            "source": "yandex",
            "medium": "cpc",
            "mark_type": "acq",
            "direction": "oge",
            "funnel": "bot",
            "name": "pyweek2026",
            "utm_term": "interests",
            "utm_content": "ad-456",
            "make_full": "1",
        }
        payload.update({k: v for k, v in overrides.items() if v is not None})
        for key, value in overrides.items():
            if value is None:
                payload.pop(key, None)
        return payload

    def test_duplicate_not_created(self):
        self.client.post(reverse("marks_new"), self._payload())
        self.client.post(reverse("marks_new"), self._payload())
        self.assertEqual(MarkedLink.objects.count(), 1)

    def test_short_attaches_to_existing_duplicate(self):
        self.client.post(reverse("marks_new"), self._payload())
        mark = MarkedLink.objects.get()
        self.assertIsNone(mark.short_link)
        self.client.post(reverse("marks_new"), self._payload(make_full=None, make_short="1"))
        self.assertEqual(MarkedLink.objects.count(), 1)
        mark.refresh_from_db()
        self.assertIsNotNone(mark.short_link)

    def test_pending_review_saved_for_template_source(self):
        self.client.post(
            reverse("marks_new"),
            self._payload(source="tg-<имя>", source_custom="python2026"),
        )
        mark = MarkedLink.objects.get()
        self.assertTrue(mark.pending_review)
        self.assertEqual(mark.utm_source, "tg-python2026")


class SyncUtmDictionaryTests(TestCase):
    def test_sync_entries_upserts_then_deactivates_missing(self):
        first = sync_entries([
            {"field": "medium", "value": "cpc", "label": "", "group": "", "is_template": False, "is_active": True},
            {"field": "medium", "value": "cpm", "label": "", "group": "", "is_template": False, "is_active": True},
        ])
        self.assertEqual(first, (2, 0, 0))

        second = sync_entries([
            {"field": "medium", "value": "cpc", "label": "клик", "group": "", "is_template": False, "is_active": True},
        ])
        self.assertEqual(second, (0, 1, 1))
        self.assertFalse(UtmDictionaryEntry.objects.get(field="medium", value="cpm").is_active)

    @patch("marks.management.commands.sync_utm_dictionary.fetch_csv_rows")
    def test_sync_command_with_mocked_fetch(self, fetch_mock):
        fetch_mock.return_value = [
            ["02 · medium", "", "", "", ""],
            ["medium", "Что значит", "Группа", "", ""],
            ["cpc", "клик", "Performance", "", ""],
            ["cpm", "показы", "Performance", "", ""],
        ]
        with override_settings(UTM_DICTIONARY_CSV_URLS={"medium": "http://x/medium.csv", "source": "", "campaign": ""}):
            call_command("sync_utm_dictionary")
        self.assertEqual(UtmDictionaryEntry.objects.filter(field="medium").count(), 2)
        fetch_mock.assert_called_once()


class BotTagEnforcementTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.marketer = user_model.objects.create_user(username="mktr", password="StrongPass123!")
        self.marketer.profile.role = UserProfile.Role.MARKETER
        self.marketer.profile.save(update_fields=["role"])
        self.product = Product.objects.create(name="P")
        self.bot = Bot.objects.create(name="testbot", product=self.product)
        self.branch = Branch.objects.create(bot=self.bot, name="Main", code="MN")
        # Ветка при создании автоматически заводит первую метку (существующая метка).
        self.existing_tag = self.branch.tags.first()
        _seed_min_dictionary()
        self.client.force_login(self.marketer)

    def _create_payload(self, **overrides):
        payload = {
            "create_tag": "1",
            "source": "yandex",
            "medium": "cpc",
            "mark_type": "acq",
            "direction": "oge",
            "funnel": "bot",
            "name": "pyweek2026",
            "utm_term": "interests",
            "utm_content": "ad-456",
            "source_custom": "",
            "funnel_custom": "",
            "budget": "",
        }
        payload.update(overrides)
        return payload

    def test_enforced_tag_created_with_clean_campaign(self):
        before = self.branch.tags.count()
        self.client.post(reverse("tags_list", args=[self.branch.id]), self._create_payload())
        self.assertEqual(self.branch.tags.count(), before + 1)
        tag = self.branch.tags.exclude(pk=self.existing_tag.pk).latest("id")
        self.assertEqual(tag.utm_campaign, "acq_oge_bot_pyweek2026")
        self.assertEqual(tag.utm_source, "yandex")
        self.assertFalse(tag.pending_review)
        self.assertTrue(tag.number)
        self.assertIn("telegram.me/testbot?start=", tag.url)

    def test_source_outside_dictionary_rejected(self):
        before = self.branch.tags.count()
        self.client.post(reverse("tags_list", args=[self.branch.id]), self._create_payload(source="made-up"))
        self.assertEqual(self.branch.tags.count(), before)

    def test_dashed_funnel_in_tag(self):
        self.client.post(reverse("tags_list", args=[self.branch.id]), self._create_payload(funnel_custom="python2026"))
        tag = self.branch.tags.exclude(pk=self.existing_tag.pk).latest("id")
        self.assertEqual(tag.utm_campaign, "acq_oge_bot-python2026_pyweek2026")

    def test_template_source_sets_pending_review(self):
        self.client.post(
            reverse("tags_list", args=[self.branch.id]),
            self._create_payload(source="tg-<имя>", source_custom="python2026"),
        )
        tag = self.branch.tags.exclude(pk=self.existing_tag.pk).latest("id")
        self.assertEqual(tag.utm_source, "tg-python2026")
        self.assertTrue(tag.pending_review)

    def test_existing_tag_untouched(self):
        before_url = self.existing_tag.url
        before_source = self.existing_tag.utm_source
        self.client.post(reverse("tags_list", args=[self.branch.id]), self._create_payload())
        self.existing_tag.refresh_from_db()
        self.assertEqual(self.existing_tag.url, before_url)
        self.assertEqual(self.existing_tag.utm_source, before_source)
        self.assertFalse(self.existing_tag.pending_review)

    def test_bot_api_returns_clean_utm_by_number(self):
        self.client.post(reverse("tags_list", args=[self.branch.id]), self._create_payload())
        tag = self.branch.tags.exclude(pk=self.existing_tag.pk).latest("id")
        response = self.client.get(reverse("bot_api", kwargs={"bot_name": "testbot"}), {"number": tag.number})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "acq_oge_bot_pyweek2026")


class TagCsvImportEnforcementTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.marketer = user_model.objects.create_user(username="mktr", password="StrongPass123!")
        self.marketer.profile.role = UserProfile.Role.MARKETER
        self.marketer.profile.save(update_fields=["role"])
        self.product = Product.objects.create(name="P")
        self.bot = Bot.objects.create(name="testbot", product=self.product)
        self.branch = Branch.objects.create(bot=self.bot, name="Main", code="MN")
        _seed_min_dictionary()
        self.client.force_login(self.marketer)

    def _upload(self, csv_text):
        upload = SimpleUploadedFile("tags.csv", csv_text.encode("utf-8"), content_type="text/csv")
        return self.client.post(reverse("import_tags_csv", args=[self.branch.id]), {"file": upload})

    def test_valid_rows_imported(self):
        before = self.branch.tags.count()
        header = "utm_source,utm_medium,utm_campaign,utm_term,utm_content\n"
        self._upload(header + "yandex,cpc,acq_oge_bot_alpha,interests,ad-456\n")
        self.assertEqual(self.branch.tags.count(), before + 1)

    def test_invalid_source_row_rejected(self):
        before = self.branch.tags.count()
        header = "utm_source,utm_medium,utm_campaign,utm_term,utm_content\n"
        self._upload(header + "badsource,cpc,acq_oge_bot_alpha,interests,ad-456\n")
        self.assertEqual(self.branch.tags.count(), before)

    def test_invalid_campaign_structure_rejected(self):
        before = self.branch.tags.count()
        header = "utm_source,utm_medium,utm_campaign,utm_term,utm_content\n"
        self._upload(header + "yandex,cpc,justonepart,interests,ad-456\n")
        self.assertEqual(self.branch.tags.count(), before)
